import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
INDEX_PATH = BASE_DIR / "index.html"
TABLE_PATH = BASE_DIR / "designer.xlsx"


def _normalize(text: str) -> str:
    cleaned = str(text or "").strip().lower()
    if not cleaned:
        return ""
    cleaned = unicodedata.normalize("NFD", cleaned)
    cleaned = "".join(ch for ch in cleaned if unicodedata.category(ch) != "Mn")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


def _format_designer_name(name: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(name or "").strip())
    if not cleaned:
        return ""

    def format_word(word: str) -> str:
        if not word:
            return word
        if word.isupper() and len(word) <= 3:
            return word
        return word[0].upper() + word[1:].lower()

    def format_token(token: str) -> str:
        parts = token.split("-")
        return "-".join(format_word(part) for part in parts)

    return " ".join(format_token(token) for token in cleaned.split(" "))


def _load_designers() -> dict[str, str]:
    if not TABLE_PATH.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {TABLE_PATH}")

    wb = load_workbook(TABLE_PATH)
    ws = wb[wb.sheetnames[0]]

    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        designer = _format_designer_name(row[1] or "")
        produto = str(row[2] or "").strip()
        if not designer or not produto:
            continue
        mapping[_normalize(produto)] = designer
    return mapping


def _append_designer(desc: str, designer: str) -> str:
    cleaned = desc.strip()
    phrase_pattern = re.compile(
        rf"\bdesign de\s+{re.escape(designer)}\b",
        flags=re.IGNORECASE,
    )
    if phrase_pattern.search(cleaned):
        return phrase_pattern.sub(f"design de {designer}", cleaned)
    if cleaned and cleaned[-1] not in ".!?":
        cleaned += "."
    return f"{cleaned} design de {designer}."


def main() -> None:
    mapping = _load_designers()

    content = INDEX_PATH.read_text(encoding="utf-8")
    pattern = re.compile(
        r"name:\s*'(?P<name>[^']+)'\s*,\s*desc:\s*'(?P<desc>[^']*)'",
        re.DOTALL,
    )

    updated = 0

    def replacer(match: re.Match) -> str:
        nonlocal updated
        name = match.group("name")
        desc = match.group("desc")
        designer = mapping.get(_normalize(name))
        if not designer:
            return match.group(0)
        new_desc = _append_designer(desc, designer)
        if new_desc == desc:
            return match.group(0)
        updated += 1
        return match.group(0).replace(
            f"desc: '{desc}'", f"desc: '{new_desc}'"
        )

    new_content = pattern.sub(replacer, content)
    if new_content != content:
        INDEX_PATH.write_text(new_content, encoding="utf-8")

    print(f"Descricoes atualizadas: {updated}")


if __name__ == "__main__":
    main()
