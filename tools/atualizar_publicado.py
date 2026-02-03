import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
INDEX_PATH = BASE_DIR / "index.html"
TABLE_PATH = BASE_DIR / "tabela_produtos_site.xlsx"


def _published_codes_from_index() -> set[str]:
    content = INDEX_PATH.read_text(encoding="utf-8")
    codes = set(re.findall(r"code:\s*'([^']+)'", content))
    return {c.strip().upper() for c in codes if c and str(c).strip()}


def main() -> None:
    if not TABLE_PATH.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {TABLE_PATH}")

    published = _published_codes_from_index()

    wb = load_workbook(TABLE_PATH)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]

    try:
        code_col = headers.index("codigo") + 1
    except ValueError as exc:
        raise ValueError("Coluna 'codigo' nao encontrada na planilha (linha 1).") from exc

    if "publicado" in headers:
        pub_col = headers.index("publicado") + 1
    else:
        pub_col = len(headers) + 1
        ws.cell(row=1, column=pub_col).value = "publicado"

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=code_col).value
        code = str(value).strip().upper() if value is not None else ""
        ws.cell(row=row, column=pub_col).value = "SIM" if code in published else "NÃO"

    wb.save(TABLE_PATH)


if __name__ == "__main__":
    main()
